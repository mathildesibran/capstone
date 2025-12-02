# src/anomalies.py

import os
import calendar

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def auto_width_excel(path: str) -> None:
    """
    Automatically adjust column widths in an Excel file.
    """
    wb = load_workbook(path)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        column = col[0].column

        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        adjusted_width = max_length + 2
        ws.column_dimensions[get_column_letter(column)].width = adjusted_width

    wb.save(path)


# ============================================================================
# A1 — Day-of-Week Effect
# ============================================================================


def analyze_weekday_effect(df: pd.DataFrame):
    """
    Analyze the day-of-week anomaly.

    For each ticker and each weekday, this function computes:
    - Mean daily return
    - Return standard deviation
    - Probability of outperforming the market
    - Number of observations

    The results are saved both at the ticker level and in a global aggregated table.
    """
    print("\nRunning analysis A1: Day-of-Week effect")

    grouped = df.groupby(["ticker", "day_of_week"]).agg(
        mean_return=("daily_return", "mean"),
        std_return=("daily_return", "std"),
        prob_outperform=("outperform", "mean"),
        n_obs=("daily_return", "count"),
    ).reset_index()

    grouped["day_name"] = grouped["day_of_week"].map(lambda x: calendar.day_name[x])

    global_stats = grouped.groupby(["day_of_week", "day_name"]).agg(
        mean_return=("mean_return", "mean"),
        std_return=("std_return", "mean"),
        prob_outperform=("prob_outperform", "mean"),
        n_obs=("n_obs", "sum"),
    ).reset_index()

    os.makedirs("results/tables", exist_ok=True)
    detail_path = "results/tables/anomaly_A1_by_ticker.xlsx"
    global_path = "results/tables/anomaly_A1_global.xlsx"

    grouped.to_excel(detail_path, index=False)
    global_stats.to_excel(global_path, index=False)

    auto_width_excel(detail_path)
    auto_width_excel(global_path)

    print(f"A1 ticker-level results saved to: {detail_path}")
    print(f"A1 global results saved to:       {global_path}")

    return grouped, global_stats


# ============================================================================
# A2 — January Effect
# ============================================================================


def analyze_january_effect(df: pd.DataFrame):
    """
    Analyze the January anomaly.

    This function compares the performance of stocks in January against the
    performance in all other months.
    """
    print("\nRunning analysis A2: January effect")

    grouped = df.groupby(["ticker", "month"]).agg(
        mean_return=("daily_return", "mean"),
        std_return=("daily_return", "std"),
        prob_outperform=("outperform", "mean"),
        n_obs=("daily_return", "count"),
    ).reset_index()

    global_stats = grouped.copy()
    global_stats["is_january"] = (global_stats["month"] == 1).astype(int)

    global_stats = global_stats.groupby("is_january").agg(
        mean_return=("mean_return", "mean"),
        std_return=("std_return", "mean"),
        prob_outperform=("prob_outperform", "mean"),
        n_obs=("n_obs", "sum"),
    ).reset_index()

    global_stats["label"] = global_stats["is_january"].map(
        {1: "January", 0: "Other months"}
    )

    os.makedirs("results/tables", exist_ok=True)
    detail_path = "results/tables/anomaly_A2_by_ticker.xlsx"
    global_path = "results/tables/anomaly_A2_global.xlsx"

    grouped.to_excel(detail_path, index=False)
    global_stats.to_excel(global_path, index=False)

    auto_width_excel(detail_path)
    auto_width_excel(global_path)

    print(f"A2 ticker-level results saved to: {detail_path}")
    print(f"A2 global results saved to:       {global_path}")

    return grouped, global_stats
