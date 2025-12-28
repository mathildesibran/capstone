# src/anomalies.py

import os
import calendar

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ======================================================================
# Utility: auto-adjust Excel column widths
# ======================================================================

def auto_width_excel(path: str) -> None:
    """
    Auto-adjust column widths for all worksheets in an Excel file.
    This improves readability of exported result tables.
    """
    wb = load_workbook(path)

    for ws in wb.worksheets:
        for col in ws.columns:
            max_length = 0
            column = col[0].column

            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    wb.save(path)


# ======================================================================
# A1 — Day-of-Week Effect (market, ticker, sector)
# ======================================================================

def analyze_weekday_effect(df: pd.DataFrame):
    """
    Day-of-week anomaly based on next-day excess returns.

    Outputs one Excel file:
      - results/tables/anomaly_A1.xlsx
        * sheet 'by_ticker'
        * sheet 'global'
        * sheet 'by_sector'
    """
    print("\nRunning analysis A1: Day-of-Week effect")

    df = df.copy()

    # Next-day excess return: stock minus market
    df["excess_return_tomorrow"] = df["daily_return_tomorrow"] - df["market_return_tomorrow"]

    # ---------------------------------------------------------------
    # 1) Ticker-level table (including sector)
    # ---------------------------------------------------------------
    grouped = (
        df.groupby(["ticker", "sector", "day_of_week"])
        .agg(
            mean_excess_return_tomorrow=("excess_return_tomorrow", "mean"),
            std_excess_return_tomorrow=("excess_return_tomorrow", "std"),
            prob_outperform_tomorrow=("outperform_tomorrow", "mean"),
            n_trades=("excess_return_tomorrow", "count"),
        )
        .reset_index()
    )

    grouped["day_name"] = grouped["day_of_week"].map(lambda x: calendar.day_name[int(x)])

    # ---------------------------------------------------------------
    # 2) Global table (averaged across tickers)
    # ---------------------------------------------------------------
    global_stats = (
        grouped.groupby(["day_of_week", "day_name"])
        .agg(
            mean_excess_return_tomorrow=("mean_excess_return_tomorrow", "mean"),
            std_excess_return_tomorrow=("std_excess_return_tomorrow", "mean"),
            prob_outperform_tomorrow=("prob_outperform_tomorrow", "mean"),
            n_trades=("n_trades", "sum"),
        )
        .reset_index()
    )

    # ---------------------------------------------------------------
    # 3) Sector-level table
    # ---------------------------------------------------------------
    sector_stats = (
        grouped.groupby(["sector", "day_of_week", "day_name"])
        .agg(
            mean_excess_return_tomorrow=("mean_excess_return_tomorrow", "mean"),
            std_excess_return_tomorrow=("std_excess_return_tomorrow", "mean"),
            prob_outperform_tomorrow=("prob_outperform_tomorrow", "mean"),
            n_trades=("n_trades", "sum"),
        )
        .reset_index()
    )

    # ---------------------------------------------------------------
    # 4) Export: one file, three worksheets
    # ---------------------------------------------------------------
    os.makedirs("results/tables", exist_ok=True)
    out_path = "results/tables/anomaly_A1.xlsx"

    with pd.ExcelWriter(out_path) as writer:
        grouped.to_excel(writer, sheet_name="by_ticker", index=False)
        global_stats.to_excel(writer, sheet_name="global", index=False)
        sector_stats.to_excel(writer, sheet_name="by_sector", index=False)

    auto_width_excel(out_path)
    print(f"A1 results saved to: {out_path}")

    return grouped, global_stats, sector_stats


# ======================================================================
# A2 — January Effect
# ======================================================================

def analyze_january_effect(df: pd.DataFrame):
    """
    January anomaly based on next-day excess returns.

    Outputs one Excel file:
      - results/tables/anomaly_A2.xlsx
        * sheet 'by_ticker'
        * sheet 'global'
        * sheet 'by_sector'
    """
    print("\nRunning analysis A2: January effect")

    df = df.copy()

    df["excess_return_tomorrow"] = df["daily_return_tomorrow"] - df["market_return_tomorrow"]

    # 1) Ticker-level table by month
    grouped = (
        df.groupby(["ticker", "month"])
        .agg(
            mean_excess_return_tomorrow=("excess_return_tomorrow", "mean"),
            std_excess_return_tomorrow=("excess_return_tomorrow", "std"),
            prob_outperform_tomorrow=("outperform_tomorrow", "mean"),
            n_trades=("excess_return_tomorrow", "count"),
        )
        .reset_index()
    )

    grouped["is_january"] = (grouped["month"] == 1).astype(int)

    # 2) Global comparison: January vs other months
    global_stats = (
        grouped.groupby("is_january")
        .agg(
            mean_excess_return_tomorrow=("mean_excess_return_tomorrow", "mean"),
            std_excess_return_tomorrow=("std_excess_return_tomorrow", "mean"),
            prob_outperform_tomorrow=("prob_outperform_tomorrow", "mean"),
            n_trades=("n_trades", "sum"),
        )
        .reset_index()
    )

    global_stats["label"] = global_stats["is_january"].map({1: "January", 0: "Other months"})

    # 3) Sector-level comparison: January vs other months
    if "is_january" not in df.columns:
        df["is_january"] = (df["month"] == 1).astype(int)

    sector_stats = (
        df.groupby(["sector", "is_january"])
        .agg(
            mean_excess_return_tomorrow=("excess_return_tomorrow", "mean"),
            std_excess_return_tomorrow=("excess_return_tomorrow", "std"),
            prob_outperform_tomorrow=("outperform_tomorrow", "mean"),
            n_trades=("excess_return_tomorrow", "count"),
        )
        .reset_index()
    )

    sector_stats["label"] = sector_stats["is_january"].map({1: "January", 0: "Other months"})

    # 4) Export: one file, three worksheets
    os.makedirs("results/tables", exist_ok=True)
    out_path = "results/tables/anomaly_A2.xlsx"

    with pd.ExcelWriter(out_path) as writer:
        grouped.to_excel(writer, sheet_name="by_ticker", index=False)
        global_stats.to_excel(writer, sheet_name="global", index=False)
        sector_stats.to_excel(writer, sheet_name="by_sector", index=False)

    auto_width_excel(out_path)
    print(f"A2 results saved to: {out_path}")

    return grouped, global_stats, sector_stats


def analyze_january_effect_by_sector(df: pd.DataFrame):
    """
    Compatibility wrapper: returns the sector-level worksheet produced by A2.
    """
    _, _, sector_stats = analyze_january_effect(df)
    return sector_stats


# ======================================================================
# Generic helper for binary anomalies (0/1 indicators)
# ======================================================================

def _generic_binary_anomaly(
    df: pd.DataFrame,
    col: str,
    label_1: str,
    label_0: str,
    excel_filename: str,
):
    """
    Generic workflow for a binary anomaly indicator:

    - Compute next-day excess returns
    - Aggregate statistics:
        * by ticker (with sector)
        * global
        * by sector
    - Export one Excel file with three worksheets
    """
    df = df.copy()

    df["excess_return_tomorrow"] = df["daily_return_tomorrow"] - df["market_return_tomorrow"]

    # 1) Ticker-level table (including sector)
    grouped = (
        df.groupby(["ticker", "sector", col])
        .agg(
            mean_excess_return_tomorrow=("excess_return_tomorrow", "mean"),
            std_excess_return_tomorrow=("excess_return_tomorrow", "std"),
            prob_outperform_tomorrow=("outperform_tomorrow", "mean"),
            n_trades=("excess_return_tomorrow", "count"),
        )
        .reset_index()
    )

    # 2) Global table (averaged across tickers)
    global_stats = (
        grouped.groupby(col)
        .agg(
            mean_excess_return_tomorrow=("mean_excess_return_tomorrow", "mean"),
            std_excess_return_tomorrow=("std_excess_return_tomorrow", "mean"),
            prob_outperform_tomorrow=("prob_outperform_tomorrow", "mean"),
            n_trades=("n_trades", "sum"),
        )
        .reset_index()
    )
    global_stats["label"] = global_stats[col].map({1: label_1, 0: label_0})

    # 3) Sector-level table
    sector_stats = (
        grouped.groupby(["sector", col])
        .agg(
            mean_excess_return_tomorrow=("mean_excess_return_tomorrow", "mean"),
            std_excess_return_tomorrow=("std_excess_return_tomorrow", "mean"),
            prob_outperform_tomorrow=("prob_outperform_tomorrow", "mean"),
            n_trades=("n_trades", "sum"),
        )
        .reset_index()
    )
    sector_stats["label"] = sector_stats[col].map({1: label_1, 0: label_0})

    # 4) Export: one file, three worksheets
    os.makedirs("results/tables", exist_ok=True)
    out_path = f"results/tables/{excel_filename}"

    with pd.ExcelWriter(out_path) as writer:
        grouped.to_excel(writer, sheet_name="by_ticker", index=False)
        global_stats.to_excel(writer, sheet_name="global", index=False)
        sector_stats.to_excel(writer, sheet_name="by_sector", index=False)

    auto_width_excel(out_path)
    print(f"{excel_filename} saved with sheets: by_ticker, global, by_sector")

    return grouped, global_stats, sector_stats


# ======================================================================
# A3 — Turn-of-the-Month Effect
# ======================================================================

def analyze_turn_of_month_effect(df: pd.DataFrame):
    """
    Turn-of-the-month anomaly (A3):
    - One Excel file: anomaly_A3.xlsx (by_ticker, global, by_sector)
    """
    print("\nRunning analysis A3: Turn-of-the-month effect")
    return _generic_binary_anomaly(
        df,
        col="turn_of_month",
        label_1="Turn-of-month days",
        label_0="Other days",
        excel_filename="anomaly_A3.xlsx",
    )


# ======================================================================
# A4 — Sell-in-May Effect
# ======================================================================

def analyze_sell_in_may_effect(df: pd.DataFrame):
    """
    Sell-in-May anomaly (A4):
    - One Excel file: anomaly_A4.xlsx (by_ticker, global, by_sector)
    """
    print("\nRunning analysis A4: Sell-in-May effect")
    return _generic_binary_anomaly(
        df,
        col="sell_in_may",
        label_1="May–October",
        label_0="November–April",
        excel_filename="anomaly_A4.xlsx",
    )


# ======================================================================
# A5 — Pre-Holiday Effect
# ======================================================================

def analyze_pre_holiday_effect(df: pd.DataFrame):
    """
    Pre-holiday anomaly (A5):
    - One Excel file: anomaly_A5.xlsx (by_ticker, global, by_sector)
    """
    print("\nRunning analysis A5: Pre-holiday effect")
    return _generic_binary_anomaly(
        df,
        col="pre_holiday",
        label_1="Pre-holiday days",
        label_0="Other days",
        excel_filename="anomaly_A5.xlsx",
    )
